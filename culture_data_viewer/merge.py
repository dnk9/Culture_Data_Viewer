import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import Dict
import logging


logging.basicConfig(level=logging.DEBUG)


def open_offline_file(offline_file: str, sheet_name: str = None) -> Dict[str, pd.DataFrame]:
    """
    Opens the "offline_file" and outputs a dictionary of dataframes: one for each reactor/condition
    Data coming from different reactors are all stored one below the other.
    """
    logging.debug(f'Running MERGE.open_offline_file with: '
                  f'\n offline_file: {offline_file} \n sheet_name: {sheet_name} ')

    # Checking if the xlsx file is fine: does it have the correct sheet?
    if sheet_name is None:
        logging.debug(f'No sheet name was provided. Importing the first sheet of the xlsx file.')
        # If no sheet name is specified, take just the first sheet in the work book
        try:
            offline_df = pd.read_excel(offline_file, sheet_name=0).dropna(subset=["Time (h)"])
        except KeyError:
            logging.error(f"'Time (h)' is a column necessary but not found in the first sheet. If the file contains"
                          f" multiple sheets, please check that the correct sheet is specified.")
            raise SystemExit(1)

    elif sheet_name in load_workbook(filename=offline_file).sheetnames:
        logging.debug(f'Loading {sheet_name} from xlsx file.')
        # This may be a bit redundant: opening the file twice
        try:
            offline_df = pd.read_excel(offline_file, sheet_name=sheet_name).dropna(subset=["Time (h)"])
        except KeyError:
            logging.error(f"'Time (h)' is a column necessary but not found in {sheet_name}.")
            raise SystemExit(1)

    else:
        logging.warning(f'Opening of offline file failed: no sheet {sheet_name} in {offline_file}', exc_info=True)
        return None

    offline_df['InoculationTime []'] = pd.to_timedelta(offline_df["Time (h)"], unit="h")
    offline_df["Sample"] = True

    logging.debug(f'Reactors data found in the current offline file: {offline_df["Condition"].unique()}')
    offline_df_dict = dict(tuple(offline_df.groupby("Condition")))
    return offline_df_dict


def merge_cdf_to_offline_df(cdf: pd.DataFrame, offline_df: pd.DataFrame) -> pd.DataFrame:
    """
    merges a single cdf to a single offline_df. The join is a left join. Using merge_asof. Not adding new
    keys to the left dataframe.
    Implies that all the timings are correct and inside the left dataframe
    """
    logging.debug(f'Running function *MERGE.merge_cdf_to_offline_df*')

    merged = pd.merge_asof(cdf, offline_df, tolerance=pd.Timedelta('1min'), on='InoculationTime []')

    # Finding the duplicates
    duplicated = merged.duplicated(subset='Time (h)')

    # Removing the duplication, by setting back the offline_df data to NaN
    merged.loc[duplicated, offline_df.columns.drop('InoculationTime []')] = np.NaN

    # Setting the "Sample" status to False for each non-sample
    merged["Sample"].fillna(False, inplace=True)

    return merged


def merge_df_dicts(cdf_dict: Dict[str, pd.DataFrame],
                   offline_df_dict: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Merge each cdf to the corresponding offline_df.

    Use "merge_cdf_to_offline_df" for each culture in the cdf_dict and offline_df_dict. The first six
    columns in the offline_df are ignored, as they represent information about the experimental conditions.
    """
    logging.debug(f'Running *MERGE.merge_df_dicts*')
    logging.debug(f'Keys of the offline_df_dict: {offline_df_dict.keys()}')

    merged_tracks_dict = {}

    for reactor_n, cdf in cdf_dict.items():
        # Debug the problem when reactor is specified as condition in the offline file
        logging.debug(f'Reactor number in the cdf: {reactor_n}')

        # Workaround: in case "conditions" in the offline_df are expressed with str instead of int
        reactor_letters = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H"}

        try:
            offline_df = offline_df_dict[reactor_n]

        except KeyError:
            reactor_l = reactor_letters[reactor_n]
            offline_df = offline_df_dict[reactor_l]

        merged_tracks_dict[reactor_n] = merge_cdf_to_offline_df(cdf, offline_df)

    logging.debug(f'Finished running *MERGE.merge_df_dicts* \n\n')
    return merged_tracks_dict
